import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break


def transform_to_vertical_format(input_path: str, output_path: str, custom_text: str = None,
                                 columns_to_keep: list = None, empty_rows_count: int = 6) -> None:
    # Load and preprocess data
    df = pd.read_excel(input_path, usecols=columns_to_keep).dropna(how="all")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Set the sheet direction to Right-to-Left
    ws.sheet_view.rightToLeft = True

    # Predefined styles
    DEFAULT_FONT = Font(name="Calibri", size=14)
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", readingOrder=2)
    DASHED_LEFT_BORDER = Border(right=Side(border_style="dashDot", color="000000"))

    # Set page size to A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Set page margins
    ws.page_margins = PageMargins(top=0.15, bottom=0.15, right=0.15, left=0.15, header=0, footer=0)

    empty_rows = 0
    # Process each row
    for index, row in df.iterrows():
        # Track empty columns for the current row
        empty_columns = []

        # Check if any value is NaN or empty in the row and track empty columns
        for col_name, value in row.items():
            if pd.isna(value):
                empty_columns.append(col_name)

        if empty_columns:
            # Print the row and the columns that are empty
            print(f"Row {index + 2} contains empty cells in columns: {', '.join(empty_columns)}, skipping...")
            continue

        # Add لینک row only if custom_text is provided
        if custom_text:
            ws.append(["لینک:", custom_text])
            for col in [1, 2]:
                ws.cell(row=ws.max_row, column=col).font = DEFAULT_FONT
                ws.cell(row=ws.max_row, column=col).alignment = RIGHT_ALIGN
            ws.cell(row=ws.max_row, column=2).border = DASHED_LEFT_BORDER

        # Process each column with swapped data
        for col_name, value in row.items():
            ws.append([f"{col_name}:", custom_text if pd.isna(value) else value])
            for col in [1, 2]:
                ws.cell(row=ws.max_row, column=col).font = DEFAULT_FONT
                ws.cell(row=ws.max_row, column=col).alignment = RIGHT_ALIGN
            ws.cell(row=ws.max_row, column=2).border = DASHED_LEFT_BORDER

        # Add separator row based on empty_rows_count
        if empty_rows == empty_rows_count:
            ws.row_dimensions[ws.max_row + 1].page_break = True
            empty_rows = 0
        else:
            ws.append([])
            ws.append([])
            empty_rows += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save workbook
    wb.save(output_path)


if __name__ == "__main__":
    transform_to_vertical_format(
        input_path="input.xlsx",
        output_path="transformed_output.xlsx",
        custom_text="payroll.digital.gov.krd",
        columns_to_keep=["یەکەی ژمێریاری", "ئيمەیڵ", "پاسوۆرد", "ڕۆڵ"],
        empty_rows_count=6
    )
    print("Transformation complete.")
