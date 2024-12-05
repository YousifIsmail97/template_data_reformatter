import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from typing import List, Optional


def process_excel(
        input_path: str,
        output_path: str,
        columns_to_keep: List[str],
        custom_text: Optional[str] = None
) -> None:
    """
    Process Excel file with optimized formatting and structure.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output Excel file
        columns_to_keep: List of columns to retain
        custom_text: Optional text to add above headers
    """
    # Step 1: Load and preprocess data
    df = pd.read_excel(input_path, usecols=columns_to_keep)
    df.dropna(how="all", inplace=True)

    # Step 2: Prepare data structure more efficiently
    headers = df.columns.tolist()
    num_cols = len(headers)

    # Build data structure efficiently
    data_with_headers = []
    for row in df.itertuples(index=False):
        if custom_text:
            data_with_headers.append([custom_text] + [None] * (num_cols - 1))
        data_with_headers.append(headers)
        data_with_headers.append(list(row))
        data_with_headers.append([None] * num_cols)

    # Convert to DataFrame and save intermediate file
    pd.DataFrame(data_with_headers).to_excel(output_path, index=False, header=False)

    # Step 3: Apply formatting with optimized openpyxl handling
    wb = load_workbook(output_path)
    ws = wb.active

    # Define styles once (reusable objects)
    default_style = {
        'font': Font(name="Calibri", size=14),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }

    # Process rows in batches for better performance
    for row in ws.iter_rows():
        is_empty_row = all(cell.value is None for cell in row)
        for cell in row:
            cell.font = default_style['font']
            cell.alignment = default_style['alignment']
            if not is_empty_row:
                cell.border = default_style['border']

    # Optimize column width calculation
    column_widths = {}
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        column_widths[col_letter] = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        ) + 2

    # Apply column widths in bulk
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Handle custom text merging if present
    if custom_text:
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            if row[0].value == custom_text:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=num_cols
                )
                merged_cell = ws.cell(row=row_idx, column=1)
                merged_cell.alignment = default_style['alignment']
                merged_cell.border = default_style['border']

    wb.save(output_path)


if __name__ == "__main__":
    input_path = "input.xlsx"
    output_path = "output.xlsx"
    columns_to_keep = ["یەکەی ژمێریاری", "ئيمەیڵ", "پاسوۆرد", "ڕۆڵ"]
    custom_text = "payroll.digital.gov.krd"

    process_excel(input_path, output_path, columns_to_keep, custom_text)
    print(f"File saved to {output_path}")